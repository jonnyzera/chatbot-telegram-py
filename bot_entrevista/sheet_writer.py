# sheet_writer.py
from openpyxl import Workbook, load_workbook
from config import PLANILHA
import os
from datetime import datetime

def salvar_resposta(candidato, pergunta, resposta):
    if not os.path.exists(PLANILHA):
        wb = Workbook()
        ws = wb.active
        ws.append(["Candidato", "Pergunta", "Resposta", "Data/Hora"])
        wb.save(PLANILHA)

    wb = load_workbook(PLANILHA)
    ws = wb.active
    ws.append([candidato, pergunta, resposta, datetime.now().strftime("%d/%m/%Y %H:%M:%S")])
    wb.save(PLANILHA)
